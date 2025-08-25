import time
import traceback
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def unmerge_and_fill(excel_path, sheet_name, save_path=None):
    # 加载工作簿
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    # 获取所有合并区域（先保存，避免解除后丢失）
    merged_ranges = list(ws.merged_cells.ranges)

    # 第一步：先解除所有合并（关键！解除后单元格变为普通单元格，可赋值）
    for merged_range in merged_ranges:
        ws.unmerge_cells(merged_range.coord)  # 先解除合并

    # 第二步：再填充值（此时单元格已不是MergedCell，可修改value）
    for merged_range in merged_ranges:
        # 获取合并区域范围
        min_col, min_row, max_col, max_row = merged_range.bounds
        # 主值（合并前的第一个单元格值）
        main_value = ws.cell(row=min_row, column=min_col).value

        # 填充所有单元格
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=main_value)  # 此时可正常赋值

    # 转为DataFrame
    headers = [cell.value for cell in ws[2]]
    data = [list(row) for row in ws.iter_rows(min_row=3, values_only=True)]
    df = pd.DataFrame(data, columns=headers,dtype=object)

    # 保存文件
    if save_path:
        wb.save(save_path)
        print(f"处理完成，保存至：{save_path}")
    wb.close()
    # print(df.head())
    return df
def process_excel_files(table1_path, table2_path, output_table1_path, output_table2_path):
    # 读取表1，指定dtype为object以保持原始数据类型
    df1 = pd.read_excel(table1_path, dtype=object)

    # 读取表2，指定dtype为object以保持原始数据类型
    df2 = unmerge_and_fill(table2_path,sheet_name="抖音-华为星桥专卖店")
    # df2 = pd.read_excel(table2_path,sheet_name="抖音-华为星桥专卖店", dtype=object,header=1)

    # 检查必要字段是否存在
    required_fields1 = ["sku单号", "账单批次", "采购成本（元）", "服务费用（元）"]
    required_fields2 = ["sku单号", "账单批次—1"]

    missing1 = [f for f in required_fields1 if f not in df1.columns]
    missing2 = [f for f in required_fields2 if f not in df2.columns]

    if missing1 or missing2:
        error_msg = []
        if missing1:
            error_msg.append(f"表1缺少必要字段: {', '.join(missing1)}")
        if missing2:
            error_msg.append(f"表2缺少必要字段: {', '.join(missing2)}")
        raise ValueError("; ".join(error_msg))

    # 在表1中添加新列用于标记重复的SKU
    df1["二次登记状态"] = ""

    # 统计表2中每个sku单号的出现次数
    sku_counts = df2["sku单号"].value_counts()

    # 遍历表1垫资款的每一行
    for idx, row in df1.iterrows():
        current_sku = row["sku单号"]
        cost = row["采购成本（元）"]
        # 移除可能的货币符号和空格，再转换为浮点数
        cost_clean = str(cost).replace('¥', '').replace(' ', '').strip()
        cost_num = float(cost_clean)


        # print(current_sku)
        # 检查当前sku在表2中的出现次数
        count = sku_counts.get(current_sku, 0)
        # print(count)
        if count == 0:
            # 没有匹配的sku
            df1.at[idx, "二次登记状态"] = "未匹配_未找到匹配"
            continue
        elif count == 2:
            # 有两个或更多匹配的sku，做特殊标记
            df1.at[idx, "二次登记状态"] = "两个单号"
            mask = df2["sku单号"] == current_sku
            matching_rows = df2[mask]  # 获取所有匹配的行,一般是两个，一个购买一个退货
            if cost_num > 0:
                print(f"行{idx + 2}：采购成本为正数（{cost_num}）")
                filtered_rows = matching_rows[matching_rows["订单金额"].astype(float) > 0]
                first_match_idx = filtered_rows.index[0]

            elif cost_num < 0:
                print(f"行{idx + 2}：采购成本为负数（{cost_num}）")
                filtered_rows = matching_rows[matching_rows["订单金额"].astype(float) < 0]
                first_match_idx = filtered_rows.index[0]
            else:
                print(f"行{idx + 2}：采购成本为零")
                df1.at[idx, "二次登记状态"] = "未匹配_采购成本为零"
                continue
        elif count > 2:
            # 选项过多，无法排除
            df1.at[idx, "二次登记状态"] = "未匹配_匹配过多，无法排除"
            continue
        else:
            # 只有一个匹配的sku
            df1.at[idx, "二次登记状态"] = "正常匹配"
            mask = df2["sku单号"] == current_sku
            first_match_idx = df2[mask].index[0]
            #
            # dengji_num =  df2.loc[first_match_idx]["订单金额"]
            # time.sleep(500)
            #

        # 填充数据到表2
        df2.at[first_match_idx, "账单批次—1"] = row["账单批次"]
        df2.at[first_match_idx, "行类型—1"] = row["行类型"]
        df2.at[first_match_idx, "订单应付金额（元）—1"] = row["订单应付金额（元）"]
        df2.at[first_match_idx, "政府补贴（元）—1"] = row["政府补贴（元）"]
        df2.at[first_match_idx, "店铺补贴（元）—1"] = row["店铺补贴（元）"]
        df2.at[first_match_idx, "自营补贴（元）—1"] = row["自营补贴（元）"]
        df2.at[first_match_idx, "分账金额（元）—1"] = row["分账金额（元）"]
        df2.at[first_match_idx, "服务费用（元）—1"] = row["服务费用（元）"]
        df2.at[first_match_idx, "平台折扣（元）—1"] = row["平台折扣（元）"]
        df2.at[first_match_idx, "订单实付（元）—1"] = row["订单实付（元）"]
        df2.at[first_match_idx, "采购折扣比例—1"] = row["采购折扣比例"]
        df2.at[first_match_idx, "采购折扣金额（元）—1"] = row["采购折扣金额（元）"]
        df2.at[first_match_idx, "采购成本（元）—1"] = row["采购成本（元）"]
        df2.at[first_match_idx, "结算金额（元）—1"] = row["结算金额（元）"]
        df2.at[first_match_idx, "创建时间—1"] = row["创建时间"]
        df2.at[first_match_idx, "备注—1"] = row["备注"]
    '''
    店铺主体	账单批次	sku单号	订单应付金额（元）	政府补贴（元）	分账金额（元）	服务费用（元）	订单实付（元）	采购折扣比例	
    采购折扣金额（元）	采购成本（元）	结算金额（元）	创建时间	备注	行类型	店铺名

    
    账单批次—1	行类型—1	订单应付金额（元）—1	政府补贴（元）—1	店铺补贴（元）—1	自营补贴（元）—1	分账金额（元）—1	服务费用（元）—1	
    平台折扣（元）—1	订单实付（元）—1	采购折扣比例—1	采购折扣金额（元）—1	采购成本（元）—1	结算金额（元）—1	创建时间—1	备注—1
    '''

    df1.to_excel(output_table1_path, index=False, engine='openpyxl')

    # 改进的表2保存方法，避免string_conversion参数问题
    # 使用openpyxl直接创建工作簿并写入数据
    wb = Workbook()
    ws = wb.active

    # 写入表头
    for col_idx, column in enumerate(df2.columns, 1):
        ws.cell(row=1, column=col_idx, value=column)

    # 写入数据行，确保数据类型不变
    for row_idx, row in enumerate(dataframe_to_rows(df2, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            # 直接写入原始值，不进行类型转换
            ws.cell(row=row_idx, column=col_idx, value=value)

    batch_col_idx = None
    shop_col_idx = None

    # 遍历表头查找列索引
    for col_idx, column in enumerate(df2.columns, 1):
        if column == '账单批次':
            batch_col_idx = col_idx
        elif column == '店铺主体':
            shop_col_idx = col_idx

    if not batch_col_idx or not shop_col_idx:
        raise ValueError("表2中未找到'账单批次—1'或'店铺主体'列")

    # 从第二行开始处理数据（跳过标题行）
    row = 3
    while row <= ws.max_row:
        # 获取当前行的账单批次值
        current_batch = ws.cell(row=row, column=batch_col_idx).value
        if current_batch is None:
            row += 1
            continue

        # 查找连续相同的账单批次
        merge_rows = 1
        next_row = row + 1
        while next_row <= ws.max_row and ws.cell(row=next_row, column=batch_col_idx).value == current_batch:
            merge_rows += 1
            next_row += 1

        # 合并店铺主体列（只合并店铺主体，不合并账单批次）
        if merge_rows > 1:
            # 合并对应的店铺主体列
            ws.merge_cells(start_row=row, start_column=shop_col_idx,
                           end_row=row + merge_rows - 1, end_column=shop_col_idx)

            # 设置合并后单元格的对齐方式为居中
            ws.cell(row=row, column=shop_col_idx).alignment = Alignment(horizontal='center', vertical='center')

        # 处理下一组数据
        row += merge_rows


    wb.save(output_table2_path)

    print(f"处理完成！")
    print(f"已标记的表1已保存至: {output_table1_path}")
    print(f"已更新的表2已保存至: {output_table2_path}")

# 使用示例
if __name__ == "__main__":
    # 请替换为实际的文件路径
    table1_path = "./中间文件—可忽略/垫资款结果_未处理.xlsx"
    table2_path = "国补表.xlsx"
    output_table1_path = "垫资款_已标记.xlsx"
    output_table2_path = "国补_已更新.xlsx"

    try:
        process_excel_files(table1_path, table2_path, output_table1_path, output_table2_path)
    except Exception as e:
        print(f"操作失败: {str(e)}")
        traceback.print_exc()
