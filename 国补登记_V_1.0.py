import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime, time
import sys
import io
import time
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
# 开始处理3c商品名数据
def process_order_numbers(input_path):
    """处理网店单号，返回处理后的DataFrame（不保存单个文件）"""
    if not os.path.exists(input_path):
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    # 读取Excel文件
    try:
        excel_data = pd.read_excel(input_path, sheet_name=None,dtype=object)
        main_sheet = next(iter(excel_data.keys()))  # 取第一个sheet
        df = excel_data[main_sheet].copy()
    except Exception as e:
        raise Exception(f"读取Excel失败: {str(e)}")

    # 检查必要列
    if "网店单号" not in df.columns:
        raise ValueError("Excel中缺少'网店单号'列，请检查列名")

    # 定义去除末尾字母的函数
    def remove_suffix(s):
        if pd.isna(s):
            return s
        s_str = str(s)
        cleaned = re.sub(r'[A-Za-z]+$', '', s_str)  # 仅去除尾部字母
        return cleaned if cleaned else s_str

    # 添加"去后缀"列并调整位置
    order_col_idx = df.columns.get_loc("网店单号")
    df["网店单号-去后缀"] = df["网店单号"].apply(remove_suffix)
    cols = list(df.columns)
    cols.insert(order_col_idx + 1, cols.pop(cols.index("网店单号-去后缀")))  # 移到网店单号后
    df = df.reindex(columns=cols)

    print(f"📊 已处理文件: {os.path.basename(input_path)}，记录数：{len(df)}")
    return df  # 仅返回处理后的数据框

def create_summary_file( all_dataframes):
    """创建汇总表"""
    try:
        # 合并所有数据
        summary_df = pd.concat(all_dataframes, ignore_index=True)


        # 生成带时间戳的汇总文件名，避免重复
        summary_path = f"./中间文件—可忽略/网店单号汇总表.xlsx"

        # 保存汇总表
        with pd.ExcelWriter(summary_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name="汇总数据", index=False)

        print(f"\n📑 汇总表已生成，共 {len(summary_df)} 条记录")
        print(f"📌 汇总表路径：{summary_path}")
        return summary_path
    except Exception as e:
        print(f"❌ 生成汇总表失败: {str(e)}")
        return None

def batch_process_excel(input_dir):
    """
    批量处理文件夹中的所有Excel文件，仅生成汇总表

    参数:
        input_dir: 包含Excel文件的输入文件夹路径

    """
    # 验证输入文件夹
    if not os.path.isdir(input_dir):
        raise NotADirectoryError(f"输入路径不是有效的文件夹: {input_dir}")

    # 记录处理结果和所有数据
    processed_count = 0
    error_files = []
    all_data = []  # 用于存储所有处理后的数据

    # 遍历所有文件和子文件夹
    for root, dirs, files in os.walk(input_dir):
        # 处理当前目录下的Excel文件
        for file in files:
            # 只处理Excel文件
            if file.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                input_path = os.path.join(root, file)

                try:
                    # 调用处理函数，获取处理后的数据
                    processed_df = process_order_numbers(input_path)

                    # 添加来源文件信息
                    processed_df['来源文件'] = file
                    processed_df['来源路径'] = os.path.relpath(input_path, input_dir)

                    all_data.append(processed_df)
                    processed_count += 1
                except Exception as e:
                    print(f"❌ 处理失败 {file}: {str(e)}")
                    error_files.append((file, str(e)))

    # 生成汇总表
    if all_data:
        create_summary_file(all_data)
    else:
        print("\n⚠️ 没有可汇总的数据，未生成汇总表")

    # 输出处理 summary
    print("\n" + "=" * 50)
    print(f"处理完成 | 成功: {processed_count} | 失败: {len(error_files)}")

    if error_files:
        print("\n失败文件列表:")
        for file, error in error_files:
            print(f"- {file}: {error}")

    return

# 开始处理抖音店铺文件
    # 根据sku订单号去重，（弃用） 会删除退货的订单导致数据错误（退货的和购买的是同一个订单号）
def merge_excel_with_duplicates(input_dir, order_column, output_path=None):
    """
    合并Excel文件并处理重复订单，强制订单号为文本格式，新增“店铺主体”字段
    """
    # 验证输入文件夹
    if not os.path.isdir(input_dir):
        raise NotADirectoryError(f"输入路径不是有效文件夹: {input_dir}")

    # 读取所有Excel文件（强制订单列为字符串）
    all_data = []
    processed_files = []
    error_files = []

    for root, _, files in os.walk(input_dir):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                file_path = os.path.join(root, file)
                try:
                    # ==================== 新增：提取店铺主体 ====================
                    # 文件名格式：国补_店铺名_店铺主体_时间（按"_"拆分）
                    # 例如："国补自营壳货款_华为星桥专卖店_北京朝酿暮饮商贸有限公司_2025-08-10"
                    # 拆分后取第3个元素（索引2）作为店铺主体
                    # 拆分后取第5个元素（索引4）作为账单批次
                    file_name = os.path.splitext(file)[0]  # 去除文件后缀（如.xlsx）
                    parts = file_name.split("_")
                    if len(parts) >= 4:  # 确保格式正确
                        shop_name = parts[1]    #  店铺名在第2个位置（索引1）
                        shop_subject = parts[2]  # 店铺主体在第3个位置（索引2）
                        bill_batch = parts[4]  # 账单批次在第5个位置（索引4）
                    else:
                        shop_subject = "未知主体"  # 格式不符时的默认值
                        bill_batch = "未知主体"  # 格式不符时的默认值
                        print(f"⚠️ 文件名格式不标准 {file}，店铺主体设为'未知主体'")
                        print(f"⚠️ 文件名格式不标准 {file}，账单批次设为'未知账单批次'")
                    # ======================================================

                    # 第一步：先读取表头，确认订单列是否存在
                    df_header = pd.read_excel(file_path, nrows=0,dtype=object)
                    if order_column not in df_header.columns:
                        raise ValueError(f"文件 {file} 缺少订单字段: {order_column}")

                    # 第二步：读取完整数据，强制订单列为字符串
                    df = pd.read_excel(
                        file_path
                        ,dtype=object,
                        converters={order_column: str}
                    )

                    # ==================== 新增：添加店铺主体字段（放在最前面） ====================
                    df.insert(0, "店铺主体", shop_subject)  # 插入到第0列（最前面）
                    df.insert(1, "店铺名", shop_name)
                    df.insert(2, "账单批次", bill_batch)
                    # ======================================================================

                    # 添加来源信息
                    df['来源文件'] = file
                    df['来源路径'] = os.path.relpath(file_path, input_dir)
                    all_data.append(df)
                    processed_files.append(file)
                    print(f"✅ 已读取: {file} (记录数: {len(df)})，店铺主体: {shop_subject}")
                except Exception as e:
                    print(f"❌ 读取失败 {file}: {str(e)}")
                    error_files.append((file, str(e)))

    if not all_data:
        print("\n⚠️ 未找到可合并的有效数据")
        return None

    # 合并所有数据
    merged_df = pd.concat(all_data, ignore_index=True)
    total_records = len(merged_df)
    print(f"\n📊 总记录数: {total_records}")

    # 标记重复项（基于文本格式的订单号）
    merged_df['是否重复'] = merged_df.duplicated(subset=order_column, keep='first')

    # 分离主数据和重复数据
    main_df = merged_df[~merged_df['是否重复']].drop(columns=['是否重复'])
    duplicate_df = merged_df[merged_df['是否重复']].drop(columns=['是否重复'])

    # 处理输出路径
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"合并结果_去重_{timestamp}.xlsx"

    # 保存到Excel（确保订单号为文本格式）
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            main_df.to_excel(writer, sheet_name='主数据（去重后）', index=False)
            duplicate_df.to_excel(writer, sheet_name='重复数据备份', index=False)

            # 强制订单列为文本格式（原有逻辑保留）
            for sheet_name in ['主数据（去重后）', '重复数据备份']:
                worksheet = writer.sheets[sheet_name]
                col_idx = main_df.columns.get_loc(order_column)
                for row in range(1, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=col_idx + 1)
                    cell.number_format = '@'

        print(f"\n💾 合并完成，已保存至: {os.path.abspath(output_path)}")
        print(f"📌 主数据记录数: {len(main_df)}")
        print(f"📌 重复数据记录数: {len(duplicate_df)}")

        if error_files:
            print("\n❌ 处理失败的文件:")
            for file, err in error_files:
                print(f"- {file}: {err}")

        return output_path
    except Exception as e:
        print(f"\n❌ 保存文件失败: {str(e)}")
        return None
    # 根据文件名称中的订单批次来判断该文件是否下载重复。
    # 根据文件名中的订单批次去重

def merge_excel_by_batch(input_dir, order_column, output_path=None):
    """
    合并Excel文件，根据文件名中用"-"分割的第四个元素（订单批次）检测重复文件
    强制订单号为文本格式，新增“店铺主体”字段
    """
    # 验证输入文件夹
    if not os.path.isdir(input_dir):
        raise NotADirectoryError(f"输入路径不是有效文件夹: {input_dir}")

    # 读取所有Excel文件（强制订单列为字符串）
    all_data = []
    processed_files = []
    error_files = []
    processed_batches = set()  # 用于记录已处理的订单批次
    duplicate_files = []  # 用于记录重复批次的文件

    for root, _, files in os.walk(input_dir):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls', '.xlsm')):
                file_path = os.path.join(root, file)
                try:
                    # ==================== 提取订单批次（按"-"分割的第四个元素） ====================
                    file_name = os.path.splitext(file)[0]  # 去除文件后缀
                    # 按"-"分割文件名
                    batch_parts = file_name.split("_")

                    # 检查是否有至少4个元素（索引0-3）
                    if len(batch_parts) >= 4:
                        order_batch = batch_parts[4]  # 第五个元素（索引4）作为订单批次
                    else:
                        order_batch = "未知批次"
                        print(f"⚠️ 文件名格式不标准 {file}，无法提取订单批次，设为'未知批次'")

                    # 检查该订单批次是否已处理过
                    if order_batch in processed_batches and order_batch != "未知批次":
                        duplicate_files.append((file, order_batch))
                        print(f"⚠️ 订单批次重复，已跳过: {file}（批次: {order_batch}）")
                        continue
                    # ======================================================================

                    # ==================== 提取店铺主体等信息（按"_"拆分） ====================
                    # 文件名格式：国补_店铺名_店铺主体_时间（按"_"拆分）
                    parts = file_name.split("_")
                    if len(parts) >= 4:  # 确保格式正确
                        shop_name = parts[1]  # 店铺名在第2个位置（索引1）
                        shop_subject = parts[2]  # 店铺主体在第3个位置（索引2）
                        bill_batch = parts[4] if len(parts) > 4 else "未知"  # 账单批次
                    else:
                        shop_name = "未知店铺"
                        shop_subject = "未知主体"
                        bill_batch = "未知账单批次"
                        print(f"⚠️ 文件名格式不标准 {file}，店铺信息使用默认值")
                    # ======================================================

                    # 第一步：先读取表头，确认订单列是否存在
                    df_header = pd.read_excel(file_path, nrows=0,dtype=object)
                    if order_column not in df_header.columns:
                        raise ValueError(f"文件 {file} 缺少订单字段: {order_column}")

                    # 第二步：读取完整数据，强制订单列为字符串
                    df = pd.read_excel(
                        file_path
                        ,dtype=object
                        # converters={order_column: str}
                    )

                    # ==================== 添加店铺相关字段（放在最前面） ====================
                    df.insert(0, "店铺主体", shop_subject)
                    df.insert(1, "店铺名", shop_name)
                    df.insert(2, "账单批次", bill_batch)
                    # ======================================================================

                    # 添加来源信息
                    df['来源文件'] = file
                    df['来源路径'] = os.path.relpath(file_path, input_dir)
                    all_data.append(df)
                    processed_files.append(file)
                    processed_batches.add(order_batch)  # 记录已处理的订单批次
                    print(f"✅ 已读取: {file} (记录数: {len(df)})，批次: {order_batch}，店铺主体: {shop_subject}")
                except Exception as e:
                    print(f"❌ 读取失败 {file}: {str(e)}")
                    error_files.append((file, str(e)))

    if not all_data:
        print("\n⚠️ 未找到可合并的有效数据")
        return None

    # 合并所有数据
    merged_df = pd.concat(all_data, ignore_index=True)
    total_records = len(merged_df)
    print(f"\n📊 总记录数: {total_records}")

    # 处理输出路径
    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = f"合并结果_按批次去重_{timestamp}.xlsx"

    # 保存到Excel（确保订单号为文本格式）
    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            merged_df.to_excel(writer, sheet_name='合并数据', index=False)

            # 强制订单列为文本格式
            worksheet = writer.sheets['合并数据']
            col_idx = merged_df.columns.get_loc(order_column)
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx + 1)
                cell.number_format = '@'

        print(f"\n💾 合并完成，已保存至: {os.path.abspath(output_path)}")
        print(f"📌 处理的文件数: {len(processed_files)}")
        print(f"📌 跳过的重复批次文件数: {len(duplicate_files)}")

        if duplicate_files:
            print("\n⚠️ 重复批次的文件列表:")
            for file, batch in duplicate_files:
                print(f"- {file}（重复批次: {batch}）")
        if error_files:
            print("\n❌ 处理失败的文件:")
            for file, err in error_files:
                print(f"- {file}: {err}")

        return output_path
    except Exception as e:
        print(f"\n❌ 保存文件失败: {str(e)}")
        return None
# 开始比对并处理结果
def create_guobu_table(douyin_path, output_guobu_path):
    """从抖音订单表提取字段，创建初始国补登记结果表"""
    if not os.path.exists(douyin_path):
        raise FileNotFoundError(f"抖音订单文件不存在: {douyin_path}")

    # 关键修复1：读取时强制"sku单号"为字符串，避免长数字精度丢失
    try:
        douyin_df = pd.read_excel(
            douyin_path,
            dtype = object,
            # converters={"sku单号": str}  # 强制以字符串读取，保留原始格式
        )
        print(f"✅ 成功读取抖音订单表，共 {len(douyin_df)} 条记录")
    except Exception as e:
        raise Exception(f"读取抖音订单表失败: {str(e)}")

    # 要操作的所有字段，检查抖音汇总文件是否缺少字段
    required_fields = ["店铺主体","sku单号", "订单应付金额（元）", "政府补贴（元）", "采购成本（元）","服务费用（元）","行类型", "账单批次"]
    # 订单货款的单独表格字段
    dingdan_fields = ["账单批次","店铺主体", "费用项名称","行类型","sku单号","商品一级类目","商品信息.1","税率","订单应付金额（元）", "政府补贴（元）", "分账金额（元）", "服务费用（元）", "平台折扣（元）","订单实付（元）","采购折扣比例","采购折扣金额（元）","采购成本（元）", "结算金额（元）","创建时间","备注","店铺名"]
    # 垫资款的单独表格字段
    dianzi_fields = ["店铺主体","sku单号","账单批次","行类型","订单应付金额（元）", "政府补贴（元）", "分账金额（元）", "服务费用（元）", "平台折扣（元）", "订单实付（元）","采购折扣比例","采购折扣金额（元）","采购成本（元）", "结算金额（元）","创建时间","备注","店铺名"]

    missing_fields = [f for f in required_fields if f not in douyin_df.columns]
    if missing_fields:
        raise ValueError(f"抖音订单表缺少必要字段: {', '.join(missing_fields)}")

    dingdan_fields = douyin_df[
        (douyin_df["行类型"] == "订单货款") |  # 条件1：等于"订单货款"
        (douyin_df["行类型"] == "订单退款") |  # 条件2：等于"订单退款"
        (douyin_df["行类型"].str[:2] == "订单")  # 条件3：前两个字是"订单"
        ][dingdan_fields].copy() # 编写订单货款的国补登记结果
    lastName_idx = dingdan_fields.columns.get_loc("商品一级类目")
    dingdan_fields.insert(lastName_idx + 1, "名称", "")
    dingdan_fields.insert(lastName_idx + 2, "规格", "")
    dingdan_fields.insert(lastName_idx + 3, "3c商品名称", "")
    lastName_idx = dingdan_fields.columns.get_loc("政府补贴（元）")
    dingdan_fields.insert(lastName_idx + 1, "店铺补贴（元）", "")
    dingdan_fields.insert(lastName_idx + 1, "自营补贴（元）", "")

    dianzi_df = douyin_df[
        (douyin_df["行类型"] != "订单货款") &  # 排除"订单货款"
        (douyin_df["行类型"] != "订单退款") &  # 排除"订单退款"
        (douyin_df["行类型"].str[:2] != "订单")  # 排除前两个字是"订单"的情况
        ][dianzi_fields].copy()  #  编写垫资款的国补登记结果
    lastName_idx = dianzi_df.columns.get_loc("政府补贴（元）")
    dianzi_df.insert(lastName_idx + 1, "店铺补贴（元）", "")
    dianzi_df.insert(lastName_idx + 1, "自营补贴（元）", "")



    try:
        # 1. 处理订单货款表格（保存到output_guobu_path）
        with pd.ExcelWriter(output_guobu_path, engine='openpyxl') as writer:
            # 写入订单货款数据
            dingdan_fields.to_excel(writer, index=False, sheet_name="订单货款")

            # 设置"sku单号"列为文本格式
            worksheet = writer.sheets["订单货款"]
            if "sku单号" in dingdan_fields.columns:
                sku_col = dingdan_fields.columns.get_loc("sku单号") + 1  # Excel列从1开始
                for row in range(1, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=sku_col).number_format = "@"  # 文本格式标记

        # 2. 处理垫资款表格（保存到"垫资款结果.xlsx"）
        # 构建垫资款文件路径（与订单货款同目录）
        dianzi_path = f"./中间文件—可忽略/垫资款结果_未处理.xlsx"

        with pd.ExcelWriter(dianzi_path, engine='openpyxl') as writer:
            # 写入垫资款数据
            dianzi_df.to_excel(writer, index=False, sheet_name="垫资款")

            # 设置"sku单号"列为文本格式
            worksheet = writer.sheets["垫资款"]
            if "sku单号" in dianzi_df.columns:
                sku_col = dianzi_df.columns.get_loc("sku单号") + 1  # Excel列从1开始
                for row in range(1, worksheet.max_row + 1):
                    worksheet.cell(row=row, column=sku_col).number_format = "@"  # 文本格式标记

        # 打印结果信息
        print(f"✅ 国补登记结果已生成:")
        print(f"   - 订单货款表格: {output_guobu_path}（共 {len(dingdan_fields)} 条记录）")
        print(f"   - 垫资款表格: {dianzi_path}（共 {len(dianzi_df)} 条记录）")
        return dingdan_fields, dianzi_df  # 返回两个DataFrame供后续使用
    except Exception as e:
        raise Exception(f"保存国补登记结果失败: {str(e)}")

def fill_3c_name(guobu_path, wangdian_path):
    """匹配并填充3c商品名称"""
    if not os.path.exists(guobu_path):
        raise FileNotFoundError(f"国补登记结果文件不存在: {guobu_path}")
    if not os.path.exists(wangdian_path):
        raise FileNotFoundError(f"网店单号汇总表不存在: {wangdian_path}")

    # 关键修复3：读取国补表时，再次强制"sku单号"为字符串
    try:
        guobu_df = pd.read_excel(
            guobu_path,dtype=object,
            # converters={"sku单号": str}
        )
        print(f"\n✅ 读取国补登记结果，共 {len(guobu_df)} 条记录")
    except Exception as e:
        raise Exception(f"读取国补登记结果失败: {str(e)}")

    # 关键修复4：读取网店表时，强制"网店单号-去后缀"为字符串
    try:
        wangdian_df = pd.read_excel(
            wangdian_path,dtype=object
            # converters={"网店单号-去后缀": str}  # 强制字符串，避免精度丢失
        )
        print(f"✅ 读取网店单号汇总表，共 {len(wangdian_df)} 条记录")
    except Exception as e:
        raise Exception(f"读取网店单号汇总表失败: {str(e)}")

    wangdian_required = ["网店单号-去后缀", "商品名称"]
    missing_wangdian = [f for f in wangdian_required if f not in wangdian_df.columns]
    if missing_wangdian:
        raise ValueError(f"网店单号汇总表缺少必要字段: {', '.join(missing_wangdian)}")

    # 清理字符串（去除可能的空格、单引号等）
    def clean_str(s):
        if pd.isna(s):
            return ""
        return str(s).strip().strip("'")  # 去除前后空格和Excel文本标记'

    # 清理双方单号，确保匹配条件一致
    guobu_df["sku_clean"] = guobu_df["sku单号"].apply(clean_str)
    wangdian_df["wangdian_clean"] = wangdian_df["网店单号-去后缀"].apply(clean_str)

    # 构建去重的映射字典（保留第一个出现的商品名称）
    name_map = dict(
        wangdian_df.drop_duplicates("wangdian_clean", keep="first")[
            ["wangdian_clean", "商品名称"]
        ].values
    )
    print(f"✅ 已创建商品名称映射，共 {len(name_map)} 条唯一匹配关系")
    # 基于清理后的字段匹配
    guobu_df["3c商品名称"] = guobu_df["sku_clean"].map(name_map).fillna("未找到对应商品名，请检查3c商品名表格中是否存在")

    # 统计匹配结果
    matched_count = (guobu_df["3c商品名称"] != "未找到对应商品名，请检查3c商品名表格中是否存在").sum()
    print(f"✅ 匹配完成，成功填充 {matched_count} 条商品名称（共 {len(guobu_df)} 条记录）")

    # 保存最终结果（再次强制文本格式）
    try:
        with pd.ExcelWriter(guobu_path, engine='openpyxl') as writer:
            guobu_df.drop(columns=["sku_clean"]).to_excel(writer, index=False, sheet_name="国补登记结果")
            worksheet = writer.sheets["国补登记结果"]
            sku_col = guobu_df.columns.get_loc("sku单号") + 1
            for row in range(1, worksheet.max_row + 1):
                worksheet.cell(row=row, column=sku_col).number_format = "@"
        print(f"✅ 最终国补登记结果已更新: {guobu_path}")
        return guobu_df
    except Exception as e:
        raise Exception(f"更新国补登记结果失败: {str(e)}")

# 开始进行规格和其名称匹配
def parse_shop_to_sheet(shop_name):
    """根据店铺名映射到对应的sheet名（核心映射规则）"""
    # 这里的映射关系可根据实际需求扩展
    mapping = {
        "华为崇云专卖店": "河北崇云",
        "华为浩昌数码专卖店": "山东浩昌",
        "华为星桥专卖店": "河北星桥",
        "荣耀星桥专卖店": "河北星桥",
        "抖音-vivo平板旗舰店": "河北丽坤",
        "崇云平板旗舰店（店名:崇云卖场店）": "河北丽坤",
        "抖音-vivo丽坤专卖店 (前身店名-抖音-vivo全场景专卖店)": "河北丽坤",
        "华为智慧通达专卖店": "河北智慧"
        # 新增店铺名请在此处添加映射
    }

    # 尝试直接匹配
    if shop_name in mapping:
        return mapping[shop_name]


    # 未匹配到时返回原始名称并标记
    return f"未匹配_{shop_name}"

def convert_memory_format(memory_str):
    """
    将内存格式从 8G+256G 转换为 8GB+256GB
    """
    # 正则匹配数字+G的模式，然后在G前添加B
    pattern = re.compile(r'(\d+)G')
    converted = pattern.sub(r'\1GB', memory_str)
    return converted

def match_data(product_name):
    # 1. 提取“标准版”（匹配“标准版”“高配版”等中文版本描述）
    # 正则说明：匹配“版”字前的中文（如“标准”“高配”“Pro”等）
    version_pattern = re.compile(r'([\u4e00-\u9fa5A-Za-z]+版)')
    version_match = version_pattern.search(product_name)
    version = version_match.group(1) if version_match else ""  # 结果：标准版

    # 2. 提取“BTKR-W00”（匹配类似“XXX-XXX”的型号格式）
    # 正则说明：匹配由字母、数字、连字符组成的型号（至少3个字符）
    model_pattern = re.compile(r'([A-Za-z0-9]+-[A-Za-z0-9]+)')
    model_match = model_pattern.search(product_name)
    model = model_match.group(1) if model_match else ""  # 结果：BTKR-W00

    # 3. 提取“8G+128G”（匹配内存+存储格式）
    # 正则说明：匹配“数字G+数字G”的格式
    memory_pattern = re.compile(r'(\d+G\+\d+G)')
    memory_match = memory_pattern.search(product_name)
    memory = memory_match.group(1) if memory_match else ""  # 结果：8G+128G

    # 4. 提取“深空灰”（按空格分割取最后一个元素）
    color = product_name.split()[-1]  # 结果：深空灰

    return version, model, memory, color

    #  根据拿到的sheet名，取出规格型号及名称  字典
def generate_model_name_dict(file_path, sheet_name=None):
    """
    从企业库存数量中提取 规格型号 → 名称 映射字典
    自动适配列名：名称、规格型号

    :param file_path: 表格文件路径（Excel）
    :param sheet_name: 工作表名，默认取第一个
    :return: {规格型号: [名称1, 名称2...], ...}
    """
    # 先清理样式
    # clean_excel_styles_and_merges(file_path, "cleaned_" + file_path)


    # 1. 读取表格（跳过表头合并行，从第4行开始识别列名）
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        dtype = object,
        header=2,  # 第4行是列名行（A4:名称、C4:规格型号）
        usecols=["名称", "规格型号"],  # 只加载需要的列
        engine='openpyxl'
    )
    # 新增调试打印，查看读取到的数据
    # 2. 校验必要列
    required_cols = ["名称", "规格型号"]
    if not set(required_cols).issubset(df.columns):
        raise ValueError(f"表格缺少必要列！需包含 {required_cols}，当前列：{df.columns.tolist()}")

    # 3. 按规格型号分组，收集名称（去重+保留顺序）
    model_name_dict = {}
    for model, group in df.groupby("规格型号", sort=False):  # sort=False 保留原顺序
        unique_names = group["名称"].drop_duplicates().tolist()
        model_name_dict[model] = unique_names

    return model_name_dict
    #  主要代码，进行名称匹配
def count_unique_shops_with_sheet(sheet_file_path, guige_file_path,output_path,sheet_name=None):
    """
    统计表格中“店铺名”列的不重复值，并转换为对应的sheet名,根据sheet名，获取总字典。
    在国补登记结果表格中，进行 行遍历 ，对3c商品名称进行分析。然后在字典中匹配。

    参数:
        sheet_file_path: 表格文件路径（Excel格式）
        guige_file_path: 存放规格的表格 企业库存数量.xlsx
        sheet_name: 工作表名称，默认使用第一个工作表
    返回:
        元组 (店铺名种类数量, 店铺名与sheet名的映射字典)
    """
    # 检查文件是否存在
    if not os.path.exists(sheet_file_path):
        raise FileNotFoundError(f"文件不存在: {sheet_file_path}")

    # 读取表格数据
    if sheet_name:
        df = pd.read_excel(sheet_file_path,dtype=object, sheet_name=sheet_name)
    else:
        df = pd.read_excel(sheet_file_path,dtype=object
                           # ,converters={"sku单号": str}
                           )  # 默认读取第一个工作表
        # df = pd.read_excel("国补登记结果.xlsx")  # 默认读取第一个工作表

    # 检查是否包含“店铺名”列
    if "店铺名" not in df.columns:
        raise ValueError("表格中未找到“店铺名”列，请确认列名是否正确")

    # 提取不重复的店铺名
    unique_shops = df["店铺名"].drop_duplicates().dropna().tolist()
    count = len(unique_shops)

    # 生成店铺名→sheet名的映射字典
    shop_to_sheet = {
        shop: parse_shop_to_sheet(shop)
        for shop in unique_shops
    }
    # return count, shop_to_sheet

    print(f"表格中共有 {count} 种不同的店铺名，对应的sheet名如下：")
    print(shop_to_sheet)
    unique_sheets = list(set(shop_to_sheet.values()))
    print("\n===== 开始提取每个店铺对应的规格型号字典 =====")
    total_dict = {}  # 总字典：{sheet名: 规格型号字典, ...}
    for sheet in unique_sheets:
        try:
            print(f"正在处理sheet：{sheet}")
            model_dict = generate_model_name_dict(guige_file_path, sheet)
            total_dict[sheet] = model_dict
            print(f"  成功提取 {len(model_dict)} 个规格型号")
        except Exception as e:
            print(f"  处理sheet {sheet} 失败：{str(e)}")
            continue  # 跳过错误的sheet，继续处理其他
    print("\n===== 所有sheet处理完成 =====")
    print(total_dict)
    # total_dict 为存放所有规格的数据
    print("\n===== 开始匹配规格.......... =====")
    # 遍历每行进行处理
    for index, row in df.iterrows():
        # 1. 取出当前行的"店铺名"和"3c 商品名称"
        shop_name = row["店铺名"]
        product_name = row["3c商品名称"]
        # print(f"这是我在循环中拿到的数据：{shop_name} {product_name}")
        # 2. 这里放入你的一系列处理逻辑
        # 示例：将店铺名和商品名拼接作为结果
        pipei_sheet_name =  parse_shop_to_sheet(shop_name) # 字典匹配sheet名称

        version,model,memory,color = match_data(product_name) # 从这个里面拿规格。 四个参数分别是 版本 型号 内存 颜色
        memory = convert_memory_format(memory)  # 格式化 内存大小 8G+256G -> 8GB+256GB

        pipei_data_list  = total_dict[pipei_sheet_name].get(model, f"未匹配到该型号{model}")
        # print(pipei_data_list)
        # 筛选同时包含内存和颜色的项
        final_result = [
            item for item in pipei_data_list
            if memory in item and color in item
        ]

        if len(final_result) == 1 :

            processed_result = final_result[0]
            df["规格"] = df["规格"].astype(str)
            df.at[index, "规格"] = model
        elif len(final_result) >= 2 :

            # print(final_result)
            # 还需进一步排除
            # 定义非标准版的关键词列表（可根据实际情况扩展）
            non_standard_versions = ["柔光版", "灵动版", "Pro版", "青春版"]
            # 分情况筛选
            if version == "标准版":
                # 标准版：排除包含任何非标准版关键词的项
                final_result = [
                    item for item in final_result
                    if not any(v in item for v in non_standard_versions)
                ]
            else:
                # 其他版本：直接匹配包含该版本关键词的项
                final_result = [
                    item for item in final_result
                    if version in item
                ]
            # 处理之后，在判断以下是否拿到唯一值。
            if len(final_result) == 1:
                processed_result = final_result[0]
                df["规格"] = df["规格"].astype(str)
                df.at[index, "规格"] = model
            elif len(final_result) >= 2:
                processed_result = "无法排除到唯一值_请向工程师反馈"

        elif len(final_result) == 0 :
            if isinstance(pipei_data_list, list):
                processed_result = f"有{model}规格，但是没有对应的配置"

            elif isinstance(pipei_data_list, str):  # 判断是否为字符串 未匹配到该型号
                processed_result = pipei_data_list  # 是字符串则返回本身 未匹配到该型号

        # print(processed_result)
        # print("========================")

        # 3. 将处理结果存入当前行的"名称"列
        df["名称"] = df["名称"].astype(str)
        df.at[index, "名称"] = processed_result


    # 写入Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="数据结果", index=False)  # index=False不保存索引列


# 整理表格格式
def document_file(file_path, output_path=None, sheet_name=None):
    """
    合并表格中相同的"账单批次"及对应"店铺主体"单元格，保持"sku单号"为文本类型

    参数:
        file_path: 输入Excel文件路径
        output_path: 输出文件路径，None则覆盖原文件
        sheet_name: 工作表名称，None则使用第一个工作表
    """
    # 确定输出路径
    if output_path is None:
        output_path = file_path

    # 读取Excel文件获取工作表信息
    excel_file = pd.ExcelFile(file_path)

    # 如果未指定工作表，使用第一个工作表
    if sheet_name is None:
        sheet_name = excel_file.sheet_names[0]
        print(f"使用工作表: {sheet_name}")

    # 读取数据，确保sku单号为字符串类型
    df = pd.read_excel(file_path, sheet_name=sheet_name, dtype=object)

    # 确保数据按账单批次排序（相同的排在一起）
    df = df.sort_values(by='账单批次')

    # 创建临时文件
    temp_file = f'temp_bill_merge_{int(time.time())}.xlsx'
    df.to_excel(temp_file, sheet_name=sheet_name, index=False)

    # 使用openpyxl加载临时文件
    wb = openpyxl.load_workbook(temp_file)
    ws = wb[sheet_name]

    # 获取账单批次列、店铺主体列和sku单号列的索引
    batch_col = None
    shop_col = None
    sku_col = None
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value == '账单批次':
                batch_col = cell.column_letter
            elif cell.value == '店铺主体':
                shop_col = cell.column_letter
            elif cell.value == 'sku单号':
                sku_col = cell.column_letter

    if not batch_col or not shop_col or not sku_col:
        raise ValueError("表格中未找到'账单批次'、'店铺主体'或'sku单号'列")

    # 从第二行开始处理数据（跳过标题行）
    row = 2
    while row <= ws.max_row:
        current_batch = ws[f"{batch_col}{row}"].value
        if current_batch is None:
            row += 1
            continue

        # 查找连续相同的账单批次
        merge_rows = 1
        next_row = row + 1
        while next_row <= ws.max_row and ws[f"{batch_col}{next_row}"].value == current_batch:
            merge_rows += 1
            next_row += 1

        # 合并单元格
        if merge_rows > 1:
            # 合并账单批次列
            ws.merge_cells(f"{batch_col}{row}:{batch_col}{row + merge_rows - 1}")
            # 按相同行数合并对应的店铺主体列
            ws.merge_cells(f"{shop_col}{row}:{shop_col}{row + merge_rows - 1}")

            # 设置合并后单元格的对齐方式为居中
            ws[f"{batch_col}{row}"].alignment = Alignment(horizontal='center', vertical='center')
            ws[f"{shop_col}{row}"].alignment = Alignment(horizontal='center', vertical='center')

        row += merge_rows

    # 保存临时文件并关闭
    wb.save(temp_file)
    wb.close()

    # 处理文件替换
    try:
        # 确保目标文件未被占用
        if os.path.exists(output_path):
            os.remove(output_path)
        os.rename(temp_file, output_path)
        print(f"✅处理完成，文件已保存至: {output_path}")
    except PermissionError:
        print(f"错误: 文件 {output_path} 可能被其他程序占用，请关闭后重试")
        print(f"处理后的文件临时保存为: {temp_file}")
    except Exception as e:
        print(f"处理文件时发生错误: {str(e)}")
        if os.path.exists(temp_file):
            os.remove(temp_file)

def main(process_step):
    """
    主函数，根据传入的步骤参数执行对应流程

    参数:
        process_step: 0=全流程，1=处理3c商品表，2=处理抖音店铺文件，3=比对并生成结果
    """

    # 步骤1：处理3c商品表
    def step1():
        print("\n===== 开始执行步骤1：预处理3c商品表 =====")
        input_folder = "3c商品名表格"
        try:
            batch_process_excel(input_dir=input_folder)
            print("===== 步骤1执行完成 =====")
        except Exception as e:
            print(f"步骤1执行失败: {str(e)}")

    # 步骤2：处理抖音店铺文件
    def step2():
        print("\n===== 开始执行步骤2：预处理抖音店铺文件 =====")
        input_folder = "抖音表格"
        order_field = "sku单号"
        output_file = f"./中间文件—可忽略/抖音订单合并结果.xlsx"
        try:
            merge_excel_by_batch(
                input_dir=input_folder,
                order_column=order_field,
                output_path=output_file
            )
            print("===== 步骤2执行完成 =====")
        except Exception as e:
            print(f"步骤2执行失败: {str(e)}")

    # 步骤3：比对并生成结果
    def step3():
        print("\n===== 开始执行步骤3：比对并生成结果 =====")
        douyin_order_path = f"./中间文件—可忽略/抖音订单合并结果.xlsx"
        wangdian_summary_path = f"./中间文件—可忽略/网店单号汇总表.xlsx"
        guobu_result_path = f"./中间文件—可忽略/国补登记结果_未匹配名称.xlsx"
        try:
            create_guobu_table(douyin_order_path, guobu_result_path)
            fill_3c_name(guobu_result_path, wangdian_summary_path)
            print(f"\n🎉 步骤3执行完成！最终结果已保存至：{os.path.abspath(guobu_result_path)}")
            print("===== 步骤3执行完成 =====")
        except Exception as e:
            print(f"步骤3执行失败: {str(e)}")

    def step4():
        print("\n===== 开始执行步骤4：根据3c商品名称以及企业规格进行名称匹配 =====")
        sheet_file_path = f"./中间文件—可忽略/国补登记结果_未匹配名称.xlsx"  # 替换为你的表格路径
        guige_file_path = "企业库存数量.xlsx"
        pipei_output_path = f"./中间文件—可忽略/国补登记结果_未处理.xlsx"
        count_unique_shops_with_sheet(sheet_file_path, guige_file_path, pipei_output_path)
        print(f"\n🎉 步骤4执行完成！")


    def step5():
        print("\n===== 步骤4：整理表格格式 =====")
        document_file(f"./中间文件—可忽略/国补登记结果_未处理.xlsx","国补登记结果.xlsx")
        # document_file(f"./中间文件—可忽略/垫资款结果_未处理.xlsx","垫资款结果.xlsx")


    # 根据传入的参数执行对应流程
    if process_step == 0:
        print("===== 开始执行全流程 =====")
        step1()
        step2()
        step3()
        step4()
        step5()
        print("\n===== 全流程执行完成 =====")
    elif process_step == 1:
        step1()
    elif process_step == 2:
        step2()
    elif process_step == 3:
        step3()
    elif process_step == 4:
        step4()
    elif process_step == 5:
        step5()
    else:
        print(f"无效参数：{process_step}，请传入 0（全流程）、1、2、3、4（单步骤）")

if __name__ == "__main__":
    print("=" * 40)
    print("           🔧 数据处理工具 - 流程选择           ")
    print("=" * 40)
    print(" 请选择需要执行的操作（输入对应数字后按回车）：")
    print("-" * 40)
    print(" ➡️ [0] 执行全流程")
    print("      包含：步骤1→步骤2→步骤3（完整处理流程）")
    print("-" * 40)
    print(" 📜 [1] 仅执行——处理3c商品表")
    print("      功能：批量处理3c商品表格并生成汇总表")
    print("-" * 40)
    print(" 📜 [2] 仅执行——处理抖音店铺文件")
    print("      功能：合并抖音订单并提取店铺主体信息")
    print("-" * 40)
    print(" 📜 [3] 仅执行——比对并生成结果")
    print("      功能：匹配订单信息并生成国补登记结果")
    print("-" * 40)
    print(" 📜 [4] 仅执行——匹配名称及规格")
    print("      功能：整理国补登记结果和垫资款结果的表格格式")
    print("-" * 40)
    print(" 📜 [5] 仅执行——整理表格格式")
    print("      功能：整理国补登记结果和垫资款结果的表格格式")
    print("=" * 40)

    # 交互式获取用户输入
    while True:
        user_input = input("\n请输入选择（0-5）：")
        try:
            step = int(user_input)
            # 验证输入范围
            if 0 <= step <= 5:
                main(step)  # 执行主程序

                # 等待用户按任意键退出
                input("\n操作已完成，按任意键并回车即可退出...")
                break  # 输入有效，执行后退出循环
            else:
                print("请输入 0-5 之间的数字！")
        except ValueError:
            print("❌输入无效，请输入整数（0-5）！")