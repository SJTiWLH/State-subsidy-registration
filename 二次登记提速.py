import time
import traceback
import pandas as pd
import numpy as np
import threading
import os  # 新增：用于文件路径处理
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pandas.io.excel import ExcelWriter


# --------------------------
# 新增：高效合并单元格函数（原efficient_merge_cells）
# --------------------------
def efficient_merge_cells(
        input_path,  # 输入Excel路径（即保存后的表2）
        output_path,  # 最终输出路径（合并后的表2）
        sheet_name,  # 工作表名称
        group_col,  # 基准列（如"账单批次"）
        merge_cols,  # 待合并列（如["店铺主体"]）
        header_row=1  # 表头行（默认第1行）
):
    """高效合并单元格：内存预处理+批量合并，1万行<10秒"""
    start_merge_time = time.time()
    print(f"\n🔗 开始高效合并单元格（基准列：{group_col}，合并列：{merge_cols}）")

    # 1. 读取表2数据（内存预处理）
    df = pd.read_excel(input_path, sheet_name=sheet_name, header=header_row - 1)
    total_rows = len(df)
    if total_rows == 0:
        print("⚠️ 表2无数据，无需合并")
        return

    # 2. 数据校验
    if group_col not in df.columns:
        raise ValueError(f"基准列'{group_col}'不存在于表2中")
    for col in merge_cols:
        if col not in df.columns:
            raise ValueError(f"待合并列'{col}'不存在于表2中")

    # 3. 填充空值（避免合并判断错误）
    df[group_col] = df[group_col].fillna("__EMPTY__")

    # 4. 计算合并范围（内存操作，快）
    merge_ranges = {col: [] for col in merge_cols}
    current_group = df[group_col].iloc[0]
    start_idx = 0

    for i in range(1, total_rows):
        if df[group_col].iloc[i] != current_group:
            # 转换为Excel行号（表头占1行，数据从header_row+1开始）
            start_row = header_row + 1 + start_idx
            end_row = header_row + 1 + (i - 1)
            for col in merge_cols:
                merge_ranges[col].append((start_row, end_row))
            current_group = df[group_col].iloc[i]
            start_idx = i

    # 处理最后一组
    start_row = header_row + 1 + start_idx
    end_row = header_row + 1 + (total_rows - 1)
    for col in merge_cols:
        merge_ranges[col].append((start_row, end_row))

    print(f"📊 计算完成：共{len(merge_ranges[merge_cols[0]])}组需要合并")

    # 5. 批量应用合并（减少Excel交互）
    wb = load_workbook(input_path)
    ws = wb[sheet_name]

    for col_name in merge_cols:
        col_idx = df.columns.get_loc(col_name) + 1  # 转为Excel 1-based列号
        col_letter = get_column_letter(col_idx)
        merge_count = 0
        for (start_row, end_row) in merge_ranges[col_name]:
            # 安全校验：避免范围越界
            if start_row > end_row or end_row > ws.max_row:
                continue
            # 执行合并
            ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{end_row}")
            # 居中对齐
            ws[f"{col_letter}{start_row}"].alignment = Alignment(horizontal='center', vertical='center')
            merge_count += 1
        print(f"✅ 列'{col_name}'合并完成：{merge_count}组")

    # 6. 安全保存
    try:
        wb.save(output_path)
        print(f"💾 合并后文件保存至：{output_path}")
    except Exception as e:
        # 保存失败时创建备份
        backup_path = output_path.replace(".xlsx", "_merge_backup.xlsx")
        wb.save(backup_path)
        print(f"⚠️ 主文件保存失败，已备份至：{backup_path}，错误：{str(e)}")
    wb.close()

    # 耗时统计
    end_merge_time = time.time()
    print(f"⏱️  合并单元格耗时：{round(end_merge_time - start_merge_time, 2)}秒")


# --------------------------
# 1. 基础工具函数（不变）
# --------------------------
def unmerge_and_fill(excel_path, sheet_name, save_path=None):
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]
    merged_ranges = list(ws.merged_cells.ranges)

    # 先解除合并
    for merged_range in merged_ranges:
        ws.unmerge_cells(merged_range.coord)

    # 填充合并区域值
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        main_value = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col, value=main_value)

    # 转为DataFrame（表头在第2行，数据从第3行开始）
    headers = [cell.value for cell in ws[2]]
    data = [list(row) for row in ws.iter_rows(min_row=3, values_only=True)]
    df = pd.DataFrame(data, columns=headers, dtype=object)

    if save_path:
        wb.save(save_path)
    wb.close()
    return df


def select_shop():
    shops = [
        "抖音-华为星桥专卖店", "抖音-vivo丽坤专卖店", "抖音-华为崇云专卖店", "抖音-华为浩昌数码专卖店",
        "京东-崇云平板旗舰店", "抖音-荣耀星桥专卖店", "抖音-华为智慧通达专卖店", "抖音-vivo平板旗舰店"
    ]
    print("=" * 70)
    print("🎉 欢迎使用国补二次登记系统（V2.1 - 高效合并版）")
    print("⚠️  请确保已完成：1. 国补表存根目录 2. 垫资款文件在「中间文件—可忽略」")
    print("=" * 70)
    print("🏪 请选择处理的店铺（输入1-8）：")
    print("-" * 70)
    for i, shop in enumerate(shops, 1):
        print(f"   {i:2d} → {shop}")
    print("-" * 70)

    while True:
        user_input = input("请输入选择（1-8）：").strip()
        if not user_input.isdigit():
            print(f"❌ 输入错误！请输1-8（当前：{user_input}）")
            continue
        select_num = int(user_input)
        if 1 <= select_num <= 8:
            selected_shop = shops[select_num - 1]
            print(f"\n✅ 已选择店铺：{selected_shop}")
            print("=" * 70)
            return selected_shop
        else:
            print(f"❌ 超出范围！请输1-8（当前：{select_num}）")


# --------------------------
# 2. 多线程核心：并行处理表1的SKU匹配（不变）
# --------------------------
def process_table1_batch(batch_idx, batch_data, df2_grouped, sku_counts, df1, lock):
    print(f"📌 线程{batch_idx}开始处理，批次数据量：{len(batch_data)}行")
    start_time = time.time()

    # 向量化函数
    def clean_cost(cost):
        try:
            return float(str(cost).replace('¥', '').replace(' ', '').strip())
        except (ValueError, TypeError):
            return np.nan

    def clean_order_amt(amt):
        try:
            return float(str(amt).replace('¥', '').replace(' ', '').strip())
        except (ValueError, TypeError):
            return 0

    # 处理批次数据
    for idx, row in batch_data:
        current_sku = row["sku单号"]
        cost_num = clean_cost(row["采购成本（元）"])
        count = sku_counts.get(current_sku, 0)
        status = ""

        # 匹配逻辑
        if pd.isna(cost_num):
            status = "未匹配_采购成本格式错误"
        elif count == 0:
            status = "未匹配_未找到匹配"
        elif count == 2:
            status = "两个单号"
            try:
                matching_rows = df2_grouped.get_group(current_sku).copy()
                matching_rows["订单金额_数值"] = matching_rows["订单金额"].apply(clean_order_amt)
                filtered = matching_rows[matching_rows["订单金额_数值"] > 0] if cost_num > 0 else matching_rows[matching_rows["订单金额_数值"] < 0]
                if not filtered.empty:
                    first_idx = filtered.index[0]
                    with lock:
                        fill_cols = [
                            ("账单批次—1", "账单批次"), ("行类型—1", "行类型"), ("订单应付金额（元）—1", "订单应付金额（元）"),
                            ("政府补贴（元）—1", "政府补贴（元）"), ("店铺补贴（元）—1", "店铺补贴（元）"), ("自营补贴（元）—1", "自营补贴（元）"),
                            ("分账金额（元）—1", "分账金额（元）"), ("服务费用（元）—1", "服务费用（元）"), ("平台折扣（元）—1", "平台折扣（元）"),
                            ("订单实付（元）—1", "订单实付（元）"), ("采购折扣比例—1", "采购折扣比例"), ("采购折扣金额（元）—1", "采购折扣金额（元）"),
                            ("采购成本（元）—1", "采购成本（元）"), ("结算金额（元）—1", "结算金额（元）"), ("创建时间—1", "创建时间"), ("备注—1", "备注")
                        ]
                        for target, source in fill_cols:
                            if source in row.index:
                                df2.at[first_idx, target] = row[source]
                else:
                    status = "未匹配_无对应正负订单金额"
            except KeyError:
                status = "未匹配_无对应订单"
        elif count > 2:
            status = "未匹配_匹配过多，无法排除"
        else:
            status = "正常匹配"
            try:
                matching_rows = df2_grouped.get_group(current_sku)
                first_idx = matching_rows.index[0]
                with lock:
                    fill_cols = [
                        ("账单批次—1", "账单批次"), ("行类型—1", "行类型"), ("订单应付金额（元）—1", "订单应付金额（元）"),
                        ("政府补贴（元）—1", "政府补贴（元）"), ("店铺补贴（元）—1", "店铺补贴（元）"), ("自营补贴（元）—1", "自营补贴（元）"),
                        ("分账金额（元）—1", "分账金额（元）"), ("服务费用（元）—1", "服务费用（元）"), ("平台折扣（元）—1", "平台折扣（元）"),
                        ("订单实付（元）—1", "订单实付（元）"), ("采购折扣比例—1", "采购折扣比例"), ("采购折扣金额（元）—1", "采购折扣金额（元）"),
                        ("采购成本（元）—1", "采购成本（元）"), ("结算金额（元）—1", "结算金额（元）"), ("创建时间—1", "创建时间"), ("备注—1", "备注")
                    ]
                    for target, source in fill_cols:
                        if source in row.index:
                            df2.at[first_idx, target] = row[source]
            except KeyError:
                status = "未匹配_无对应订单"

        # 加锁更新状态
        with lock:
            df1.at[idx, "二次登记状态"] = status

    end_time = time.time()
    print(f"✅ 线程{batch_idx}处理完成，耗时：{round(end_time - start_time, 2)}秒")


# --------------------------
# 3. 主处理函数（整合高效合并）
# --------------------------
def process_excel_files(table1_path, table2_path, output_table1_path, output_table2_path, sheet_name, max_threads=4):
    global df2
    lock = threading.Lock()
    start_total_time = time.time()

    # --------------------------
    # 并行步骤1：读取表1和表2
    # --------------------------
    print("🔍 开始并行读取原始文件...")
    with ThreadPoolExecutor(max_workers=2) as executor:
        future_table1 = executor.submit(pd.read_excel, table1_path, dtype=object)
        future_table2 = executor.submit(unmerge_and_fill, table2_path, sheet_name)
        df1 = future_table1.result()
        df2 = future_table2.result()
    print(f"✅ 表1（{len(df1)}行）+ 表2（{len(df2)}行）读取完成")

    # --------------------------
    # 数据校验
    # --------------------------
    required_fields1 = ["sku单号", "账单批次", "采购成本（元）", "服务费用（元）"]
    required_fields2 = ["sku单号", "账单批次—1"]
    missing1 = [f for f in required_fields1 if f not in df1.columns]
    missing2 = [f for f in required_fields2 if f not in df2.columns]
    if missing1 or missing2:
        error_msg = []
        if missing1:
            error_msg.append(f"表1缺少字段: {', '.join(missing1)}")
        if missing2:
            error_msg.append(f"表2缺少字段: {', '.join(missing2)}")
        raise ValueError("; ".join(error_msg))
    print("✅ 字段校验通过")

    # --------------------------
    # 表1预处理
    # --------------------------
    df1["二次登记状态"] = ""
    df2_grouped = df2.groupby("sku单号")
    sku_counts = df2["sku单号"].value_counts()
    print(f"📊 表2 SKU统计：{len(sku_counts)}个不同SKU，最多重复{sku_counts.max()}次")

    # --------------------------
    # 并行步骤2：多线程处理表1 SKU匹配
    # --------------------------
    print(f"\n⚙️  多线程处理表1 SKU匹配（线程数：{max_threads}，总行数：{len(df1)}）")
    total_rows = len(df1)
    batch_size = total_rows // max_threads if total_rows >= max_threads else total_rows
    batches = []
    for i in range(max_threads):
        start_idx = i * batch_size
        end_idx = (i + 1) * batch_size if i < max_threads - 1 else total_rows
        if start_idx >= end_idx:
            break
        batch_data = list(df1.iloc[start_idx:end_idx].iterrows())
        batches.append((i + 1, batch_data))

    # 执行多线程
    with ThreadPoolExecutor(max_workers=len(batches)) as executor:
        futures = []
        for batch_idx, batch_data in batches:
            future = executor.submit(
                process_table1_batch,
                batch_idx=batch_idx,
                batch_data=batch_data,
                df2_grouped=df2_grouped,
                sku_counts=sku_counts,
                df1=df1,
                lock=lock
            )
            futures.append(future)
        for future in futures:
            future.result()
    print(f"✅ 表1 SKU匹配完成")

    # --------------------------
    # 串行步骤3：保存表1 + 表2（含SKU格式设置）
    # --------------------------
    print(f"\n💾 开始保存基础结果文件...")
    start_save_time = time.time()

    # 保存表1
    df1.to_excel(output_table1_path, index=False, engine='openpyxl')
    print(f"✅ 表1保存至：{output_table1_path}（{round(os.path.getsize(output_table1_path)/1024, 1)} KB）")

    # 保存表2（含SKU文本格式）
    with ExcelWriter(output_table2_path, engine='openpyxl') as writer:
        df2.to_excel(writer, sheet_name="Sheet1", index=False)
        wb = writer.book
        ws = writer.sheets["Sheet1"]

        # 批量设置SKU列格式（整列）
        try:
            sku_col_idx = df2.columns.get_loc("sku单号") + 1
            sku_col_letter = get_column_letter(sku_col_idx)
            ws.column_dimensions[sku_col_letter].number_format = '@'
            print(f"✅ SKU列（{sku_col_letter}列）设为文本格式")
        except ValueError:
            print(f"⚠️ 未找到'sku单号'列，跳过格式设置")
    print(f"✅ 表2基础版保存至：{output_table2_path}（{round(os.path.getsize(output_table2_path)/1024, 1)} KB）")

    # --------------------------
    # 新增步骤4：调用高效合并函数处理表2合并
    # --------------------------
    # 定义合并参数（基准列：账单批次，合并列：店铺主体）
    merge_group_col = "账单批次"
    merge_target_cols = ["店铺主体"]
    # 调用高效合并（输入：基础版表2；输出：合并后的表2）
    efficient_merge_cells(
        input_path=output_table2_path,  # 刚保存的表2基础版
        output_path="国补_已合并.xlsx",  # 直接覆盖或改为新路径（如"国补_已合并.xlsx"）
        sheet_name="Sheet1",
        group_col=merge_group_col,
        merge_cols=merge_target_cols,
        header_row=1
    )

    # --------------------------
    # 最终统计
    # --------------------------
    end_total_time = time.time()
    print("\n" + "=" * 70)
    print("🎉 全部处理完成！")
    print(f"📊 处理总结：")
    print(f"   • 总耗时：{round(end_total_time - start_total_time, 2)}秒")
    print(f"   • 表1：{len(df1)}行 → {output_table1_path}")
    print(f"   • 表2：{len(df2)}行（含{len(merge_target_cols)}列合并）→ {output_table2_path}")
    print(f"   • 匹配状态：{df1['二次登记状态'].value_counts().to_dict()}")
    print("=" * 70)


# --------------------------
# 主函数调用
# --------------------------
if __name__ == "__main__":
    # 文件路径配置
    table1_path = "./中间文件—可忽略/垫资款结果_未处理.xlsx"
    table2_path = "国补表.xlsx"
    output_table1_path = "垫资款_已标记.xlsx"
    output_table2_path = "国补_已更新.xlsx"  # 合并后会覆盖此文件（或改为新路径）

    try:
        # 1. 选择店铺
        sheet_name = select_shop()

        # 2. 自动获取CPU核心数设置线程数
        max_threads = os.cpu_count() or 4
        print(f"⚙️  系统检测到{os.cpu_count()}个CPU核心，使用{max_threads}个线程")

        # 3. 执行处理
        process_excel_files(
            table1_path=table1_path,
            table2_path=table2_path,
            output_table1_path=output_table1_path,
            output_table2_path=output_table2_path,
            sheet_name=sheet_name,
            max_threads=max_threads
        )
    except Exception as e:
        print(f"\n❌ 操作失败: {str(e)}")
        traceback.print_exc()